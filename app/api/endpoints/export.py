"""
엑셀 Export 기능
narajangteo의 핵심 장점 도입
"""

from datetime import datetime
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.logging import logger
from app.core.security import get_current_user
from app.db.models import BidAnnouncement, User
from app.db.session import get_db
from app.schemas.query import BidSource
from app.services.rate_limiter import limiter

router = APIRouter()


@router.get("/excel")
@limiter.limit("10/minute")
async def export_bids_to_excel(
    request: Request,
    importance_score: int | None = Query(default=None, ge=1, le=3, description="중요도 필터 (1-3)"),
    source: BidSource | None = Query(default=None, description="출처 필터 (G2B, Onbid)"),
    agency: str | None = Query(default=None, min_length=1, max_length=200, description="기관명 필터"),
    session: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    입찰 공고를 엑셀로 내보내기
    narajangteo 방식 응용 - 오프라인 공유/분석 용이

    - **importance_score**: 중요도 필터 (1-3)
    - **source**: 출처 필터 (G2B, Onbid)
    - **agency**: 기관명 필터 (부분 일치)
    """
    logger.info(
        f"엑셀 Export 요청: user_id={current_user.id}, filters=(importance={importance_score}, source={source})"
    )

    # 필터 조건 구성
    conditions = []
    if importance_score:
        conditions.append(BidAnnouncement.importance_score == importance_score)
    if source:
        conditions.append(BidAnnouncement.source == source.value)
    if agency:
        # SQLAlchemy 파라미터 바인딩이 SQL Injection을 방지하므로
        # 별도의 문자열 필터링은 불필요 (LIKE 패턴의 %, _ 이스케이프만 처리)
        safe_agency = agency.replace("%", r"\%").replace("_", r"\_")
        conditions.append(BidAnnouncement.agency.like(f"%{safe_agency}%"))

    # DB 조회
    query = select(BidAnnouncement).order_by(BidAnnouncement.importance_score.desc(), BidAnnouncement.created_at.desc())

    if conditions:
        query = query.where(and_(*conditions))

    result = await session.execute(query)
    bids = result.scalars().all()

    if not bids:
        raise HTTPException(status_code=404, detail="조건에 맞는 공고가 없습니다.")

    # 엑셀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "입찰공고 목록"

    # 헤더 스타일
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 헤더 작성
    headers = [
        "번호",
        "제목",
        "기관명",
        "출처",
        "마감일",
        "추정가",
        "중요도",
        "키워드",
        "상태",
        "URL",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 데이터 작성
    for idx, bid in enumerate(bids, 1):
        ws.append(
            [
                idx,
                bid.title,
                bid.agency or "미확인",
                bid.source,
                bid.deadline.strftime("%Y-%m-%d %H:%M") if bid.deadline else "미정",
                f"{int(bid.estimated_price):,}원" if bid.estimated_price else "미공개",
                "⭐" * (bid.importance_score or 1),
                ", ".join(bid.keywords_matched) if bid.keywords_matched else "",
                bid.status or "new",
                bid.url,
            ]
        )

    # 컬럼 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 파일 생성
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"bids_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_display = f"입찰공고_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    logger.info(f"엑셀 Export 완료: {len(bids)}건, filename={filename_display}")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}; filename*=UTF-8''{quote(filename_display)}"},
    )


@router.get("/priority-agencies")
@limiter.limit("10/minute")
async def export_priority_agencies_excel(
    request: Request,
    agencies: str = Query(
        ...,
        min_length=1,
        max_length=1000,
        description="콤마로 구분된 기관명 (최대 20개)",
    ),
    session: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    우선 기관 필터링 엑셀 (narajangteo orgs.txt 방식)

    - **agencies**: 콤마로 구분된 기관명 리스트 (예: "서울대병원,국립중앙의료원")
    - 최대 20개 기관까지 지정 가능
    """
    # 기관명 목록 파싱 및 검증
    priority_list = [a.strip() for a in agencies.split(",")]

    if len(priority_list) > 20:
        raise HTTPException(status_code=400, detail="최대 20개 기관까지 지정 가능합니다.")

    for agency in priority_list:
        if len(agency) < 1 or len(agency) > 100:
            raise HTTPException(status_code=400, detail=f"기관명은 1-100자 사이여야 합니다: {agency}")

    logger.info(f"우선 기관 Export: {priority_list}")

    # 우선 기관 공고만 조회
    all_bids = []
    for agency in priority_list:
        # SQLAlchemy 파라미터 바인딩 사용 (LIKE 패턴 이스케이프만 처리)
        safe_agency = agency.replace("%", r"\%").replace("_", r"\_")
        result = await session.execute(
            select(BidAnnouncement)
            .where(BidAnnouncement.agency.like(f"%{safe_agency}%"))
            .order_by(BidAnnouncement.importance_score.desc())
        )
        bids = result.scalars().all()
        all_bids.extend(bids)

    if not all_bids:
        raise HTTPException(status_code=404, detail="우선 기관 공고가 없습니다.")

    # 중복 제거
    seen_ids = set()
    unique_bids = []
    for bid in all_bids:
        if bid.id not in seen_ids:
            seen_ids.add(bid.id)
            unique_bids.append(bid)

    # 엑셀 생성 (간소화)
    wb = Workbook()
    ws = wb.active
    ws.title = "우선 기관 공고"

    ws.append(["번호", "기관명", "제목", "마감일", "중요도", "URL"])

    for idx, bid in enumerate(unique_bids, 1):
        ws.append(
            [
                idx,
                bid.agency,
                bid.title,
                bid.deadline.strftime("%Y-%m-%d") if bid.deadline else "미정",
                "⭐" * (bid.importance_score or 1),
                bid.url,
            ]
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"priority_agencies_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filename_display = f"우선기관_공고_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}; filename*=UTF-8''{quote(filename_display)}"},
    )
